import pandas as pd
import sqlite3
from openpyxl import load_workbook

#reg_map = {'us' : 4, 'af' : 5, 'cn':6, 'me':7, 'sa':8, 'la':9, 'pa':10, 'ec':11, 'eu':12, 'se':13}
reg_map = {'us' : 'D', 'af' : 'E', 'cn':'F', 'me':'G', 'sa':'H', 'la':'I', 'pa':'J', 'ec':'K', 'eu':'L', 'se':'M'}
pt_map = {'FEHC' : 28 , 'XtaxRateEmp' : 29 , 'SGMP' : 30 , 'FTPEE' : 40 , 'NEP' : 41 , 'ISPV' : 42 , 'CCS' : 43 , 'DAC' : 44 , 'FWRP' : 33 , 
        'FLWR' : 34 , 'RMDR' : 35 , 'RIPLGF' : 36 , 'FC' : 37 , 'REFOREST' : 38 , 'XtaxFrac' : 17 , 'StrUP' : 18 , 'Wreaction' : 19 , 
        'XtaxCom' : 20 , 'ICTR' : 21 , 'IOITR' : 22 , 'IWITR' : 23 , 'Ctax' : 24 , 'SGRPI' : 25 , 'ExPS' : 3 , 'LPB' : 4 , 'LPBsplit' : 5 , 
        'LPBgrant' : 6 , 'FMPLDD' : 8 , 'TOW' : 9 , 'FPGDC' : 10 , 'Lfrac' : 11 , 'SSGDR' : 12}

cid = 'XXC-902'
wb = load_workbook('scratch/e4a-game-policies.xlsx')
con = sqlite3.connect("sdg3_game.db")

def db2xlsx(cid, round, ws):
    q = "SELECT * FROM policy_decisions WHERE game_id='" + cid + "' AND round='" + str(round) + "';"
    s1 = pd.read_sql_query(q, con)
    s1_pt = s1['pol_tag'].drop_duplicates().tolist()
    ws['A1'].value = cid
    for i in range(0,320):
        pt = s1.iloc[i, 8]
        pv = s1.iloc[i, 6]
        pr = s1.iloc[i,3]
        new_row = pt_map[pt]
        new_col = reg_map[pr]
        if pt == 'FMPLDD':
            print(f'{pt} {pr} {pv} at {new_row} : {new_col}' )
            pass
        cell = new_col+str(new_row)
        ws[cell].value = pv

ws = wb['r1']  # or wb.active for first sheet
db2xlsx(cid, 1, ws)
ws = wb['r2']  # or wb.active for scond sheet
db2xlsx(cid, 2, ws)
ws = wb['r3']  # or wb.active for third sheet
db2xlsx(cid, 3, ws)

wb.save('scratch/new_file.xlsx')